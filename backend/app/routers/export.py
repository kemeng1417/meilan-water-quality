from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse, Response
from sqlalchemy.orm import Session
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
import os
import tempfile

from app.database import get_db
from app.models import TestRecord, TestDetail, SamplePoint, Indicator, StandardLimit, WaterType, Photo
from app.services.report_gen import generate_word_report, generate_html_report

router = APIRouter(prefix="/api/export", tags=["导出"])


@router.get("/{record_id}/word")
def export_word(record_id: int, db: Session = Depends(get_db)):
    record = db.query(TestRecord).filter(TestRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    buf = generate_word_report(db, record_id)
    filename = f"{record.record_no}.docx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/{record_id}/excel")
def export_excel(record_id: int, db: Session = Depends(get_db)):
    record = db.query(TestRecord).filter(TestRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")

    water_type = db.query(WaterType).filter(WaterType.id == record.water_type_id).first()
    details = db.query(TestDetail).filter(TestDetail.record_id == record_id).all()
    indicators = db.query(Indicator).order_by(Indicator.display_order).all()
    points = db.query(SamplePoint).filter(
        SamplePoint.id.in_(list(set(d.sample_point_id for d in details)))
    ).order_by(SamplePoint.area, SamplePoint.sort_order).all()

    if record.water_type_id == 4:
        limits_1 = {l.indicator_id: l for l in db.query(StandardLimit).filter(StandardLimit.water_type_id == 1).all()}
        limits_2 = {l.indicator_id: l for l in db.query(StandardLimit).filter(StandardLimit.water_type_id == 2).all()}
    else:
        limits_1 = {l.indicator_id: l for l in db.query(StandardLimit).filter(StandardLimit.water_type_id == record.water_type_id).all()}
        limits_2 = None

    matrix = {}
    for d in details:
        matrix.setdefault(d.sample_point_id, {})[d.indicator_id] = d

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "水质检测报告"

    # ── 样式 ──
    header_font = Font(name="宋体", size=10, bold=True)
    limit_font = Font(name="宋体", size=8, italic=True, color="475569")
    normal_font = Font(name="宋体", size=10)
    red_font = Font(name="宋体", size=10, color="FF0000", bold=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    limit_fill = PatternFill(start_color="F0F5FA", end_color="F0F5FA", fill_type="solid")
    pass_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    fail_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    total_cols = 2 + len(indicators)  # 序号 + 采样点 + N指标

    # ── 标题 ──
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_cols)
    c = ws.cell(row=1, column=1, value="海口美兰机场供水站水质工作反馈")
    c.font = Font(name="宋体", size=14, bold=True)
    c.alignment = Alignment(horizontal="center")

    # ── 信息行 ──
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=total_cols)
    info = f"化验日期：{record.test_date}    报告日期：{record.report_date}    执行标准：{water_type.standard_code if water_type else '—'}"
    c = ws.cell(row=2, column=1, value=info)
    c.font = Font(name="宋体", size=10)
    c.alignment = Alignment(horizontal="center")

    header_row = 4

    # ── 表头行 ──
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.cell(row=header_row, column=1, value="序号")
    ws.cell(row=header_row, column=2, value="采样点")
    for i, ind in enumerate(indicators):
        txt = f"{ind.name}\n({ind.unit})" if ind.unit else ind.name
        ws.cell(row=header_row, column=3 + i, value=txt)

    # ── 标准参考行 ──
    limit_row = header_row + 1
    for col in range(1, total_cols + 1):
        cell = ws.cell(row=limit_row, column=col)
        cell.fill = limit_fill
        cell.font = limit_font
        cell.alignment = center_align
        cell.border = thin_border

    ws.cell(row=limit_row, column=1, value="")
    ws.cell(row=limit_row, column=2, value="标准限值")
    for i, ind in enumerate(indicators):
        lim = limits_1.get(ind.id)
        lim2 = limits_2.get(ind.id) if limits_2 else None
        txt = _fmt_limit_short(ind, lim, lim2)
        ws.cell(row=limit_row, column=3 + i, value=txt)

    # ── 数据行 ──
    for r, pt in enumerate(points):
        row = limit_row + 1 + r
        for col in range(1, total_cols + 1):
            ws.cell(row=row, column=col).border = thin_border
            ws.cell(row=row, column=col).alignment = center_align

        ws.cell(row=row, column=1, value=r + 1).font = normal_font
        ws.cell(row=row, column=2, value=pt.name).font = normal_font

        for i, ind in enumerate(indicators):
            col = 3 + i
            cell = ws.cell(row=row, column=col)
            detail = matrix.get(pt.id, {}).get(ind.id)
            if detail and detail.value_text:
                cell.value = detail.value_text
                if detail.is_qualified is False:
                    cell.font = red_font
                    cell.fill = fail_fill
                elif detail.is_qualified is True:
                    cell.font = normal_font
                    cell.fill = pass_fill
                else:
                    cell.font = normal_font
            else:
                cell.value = "—"
                cell.font = normal_font

    # ── 列宽 ──
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    for i in range(len(indicators)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 13
    ws.row_dimensions[limit_row].height = 20

    # ── 异常汇总 ──
    conclusion_row = limit_row + 1 + len(points) + 1
    ws.merge_cells(start_row=conclusion_row, start_column=1, end_row=conclusion_row, end_column=total_cols)
    abnormal = [d for d in details if d.is_abnormal]
    if record.conclusion:
        conc_text = record.conclusion
    elif abnormal:
        conc_text = f"超标项: {len(abnormal)} 项 | 本次检测存在超标项目，需整改。"
    else:
        conc_text = f"结论：本次检测项目全部合格，符合{water_type.standard_code if water_type else '相关'}标准要求。"
    c = ws.cell(row=conclusion_row, column=1, value=conc_text)
    c.font = Font(name="宋体", size=10, color="FF0000" if (record.is_abnormal) else "000000")

    # ── 签名行 ──
    sig_row = conclusion_row + 1
    ws.merge_cells(start_row=sig_row, start_column=1, end_row=sig_row, end_column=total_cols)
    c = ws.cell(row=sig_row, column=1,
                value=f"化验员：{record.tester}              审核人：{record.reviewer or '___________'}              日期：{record.report_date}")
    c.font = Font(name="宋体", size=10)

    # ── 照片 ──
    photos = db.query(Photo).filter(Photo.record_id == record_id).all()
    if photos:
        photo_start = sig_row + 2
        ws.merge_cells(start_row=photo_start, start_column=1, end_row=photo_start, end_column=total_cols)
        c = ws.cell(row=photo_start, column=1, value="现场照片：")
        c.font = Font(name="宋体", size=10, bold=True)

        photo_by_point: dict[int, list] = {}
        for p in photos:
            photo_by_point.setdefault(p.sample_point_id, []).append(p)

        sp_map = {sp.id: sp for sp in points}
        row = photo_start + 1
        for sp_id, pt_photos in photo_by_point.items():
            sp_name = sp_map[sp_id].name if sp_id in sp_map else f"点位{sp_id}"
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=total_cols)
            filenames = '、'.join([p.original_name for p in pt_photos])
            c = ws.cell(row=row, column=1, value=f"  {sp_name}：{filenames}")
            c.font = Font(name="宋体", size=9)
            c.alignment = Alignment(wrap_text=True)
            row += 1

    # ── 冻结表头 ──
    ws.freeze_panes = ws.cell(row=limit_row + 1, column=3)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"{record.record_no}.xlsx"
    return StreamingResponse(buf, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/{record_id}/html")
def export_html(record_id: int, db: Session = Depends(get_db)):
    record = db.query(TestRecord).filter(TestRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    html = generate_html_report(db, record_id)
    filename = f"{record.record_no}.html"
    return Response(content=html.encode("utf-8"), media_type="text/html; charset=utf-8",
                    headers={"Content-Disposition": f"attachment; filename={filename}"})


@router.get("/{record_id}/pdf")
def export_pdf(record_id: int, db: Session = Depends(get_db)):
    record = db.query(TestRecord).filter(TestRecord.id == record_id).first()
    if not record:
        raise HTTPException(status_code=404, detail="记录不存在")
    try:
        import pythoncom
        from win32com import client as wc

        word_buf = generate_word_report(db, record_id)
        tmp_docx = os.path.join(tempfile.gettempdir(), f"_rpt_{record_id}.docx")
        tmp_pdf = os.path.join(tempfile.gettempdir(), f"_rpt_{record_id}.pdf")
        with open(tmp_docx, "wb") as f:
            f.write(word_buf.getvalue())

        pythoncom.CoInitialize()
        try:
            word = wc.DispatchEx("Word.Application")
            word.Visible = False
            word.DisplayAlerts = False
            doc = word.Documents.Open(tmp_docx)
            doc.ExportAsFixedFormat(tmp_pdf, 17)  # 17 = wdExportFormatPDF
            doc.Close(False)
            word.Quit()
        finally:
            pythoncom.CoUninitialize()

        if os.path.exists(tmp_pdf):
            with open(tmp_pdf, "rb") as f:
                pdf_bytes = f.read()
            os.unlink(tmp_docx)
            os.unlink(tmp_pdf)
            filename = f"{record.record_no}.pdf"
            return Response(content=pdf_bytes, media_type="application/pdf",
                            headers={"Content-Disposition": f"attachment; filename={filename}"})
        raise Exception("PDF generation failed")
    except Exception:
        # Fallback: return HTML (user can print to PDF from browser using print button)
        html = generate_html_report(db, record_id)
        filename = f"{record.record_no}.pdf.html"
        return Response(content=html.encode("utf-8"), media_type="text/html; charset=utf-8",
                        headers={"Content-Disposition": f"attachment; filename={filename}"})


def _fmt_limit_short(indicator, limit, limit2=None) -> str:
    def _fmt_one(lim):
        if not lim:
            return None
        if lim.qual_check:
            return lim.qual_check
        if lim.min_value is not None and lim.max_value is not None:
            return f"{lim.min_value}-{lim.max_value}"
        if lim.max_value is not None:
            return f"≤{lim.max_value}"
        if lim.min_value is not None:
            return f"≥{lim.min_value}"
        return None
    t1 = _fmt_one(limit)
    if not limit2:
        return t1 or '—'
    t2 = _fmt_one(limit2)
    if not t2 or t1 == t2:
        return t1 or '—'
    return f"出厂:{t1} 末梢:{t2}"
